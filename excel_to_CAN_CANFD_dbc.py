import openpyxl
from pathlib import Path
import sys
import re
import datetime
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
import tkinter.scrolledtext as scrolledtext

# --- Constants ---

# --- Helper Functions ---
def hex_to_dec(hex_str):
    """Converts a hex string (like '0xAA') to a decimal integer."""
    try:
        return int(str(hex_str), 16)
    except (ValueError, TypeError):
        return 0

def get_valid_dlc(input_dlc):
    """Finds the smallest valid CAN FD DLC that is >= the input DLC."""
    valid_canfd_lengths = [0, 1, 2, 3, 4, 5, 6, 7, 8, 12, 16, 20, 24, 32, 48, 64]
    if input_dlc is None:
        return 8 # Default DLC
    try:
        dlc = int(input_dlc)
        if dlc in valid_canfd_lengths:
            return dlc
        for valid_dlc in valid_canfd_lengths:
            if valid_dlc >= dlc:
                return valid_dlc
        return 64 # Max length if it exceeds all values
    except (ValueError, TypeError):
        return 8

class CANDBCGenerator:
    """
    Generates a CAN FD DBC file from a specified Excel sheet format.
    """
    def __init__(self, excel_path, sheet_name):
        self.excel_path = Path(excel_path)
        self.use_canfd = False  # Will be detected during parsing
        self.messages = {}
        self.ecus = set()
        self.ecu_addresses = {}
        self.sheet_name = sheet_name

    def validate_identifier(self, value, field_name, row_idx):
        """Validates that the value contains only alphanumeric characters and underscores."""
        if not value: return
        # DBC identifiers should strictly be alphanumeric + underscore.
        # This regex rejects spaces, commas, colons, semicolons, and other special chars.
        if not re.match(r'^[a-zA-Z0-9_]+$', str(value)):
            raise RuntimeError(
                f"데이터 오류 감지!\n"
                f"시트: '{self.sheet_name}', 위치: {row_idx}행\n"
                f"항목: {field_name}\n"
                f"값: '{value}'\n\n"
                f"원인: 공백, 특수문자(:, ;, , 등) 또는 허용되지 않는 문자가 포함되어 있습니다.\n"
                f"해당 시트의 변환을 중단합니다."
            )

    def parse_signals(self):
        """Parses messages and signals from the 'Signals' sheet of the Excel file."""
        print(f"Excel 파일: {self.excel_path}")
        print(f"Sheet 이름: {self.sheet_name}")

        try:
            workbook = openpyxl.load_workbook(self.excel_path)
            sheet = workbook[self.sheet_name]
        except Exception as e:
            print(f"오류: 엑셀 파일을 읽는 중 문제가 발생했습니다. {e}", file=sys.stderr)
            raise e

        # Find the first row that looks like a header (has multiple columns filled)
        header_row_index = 1
        for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            # 값이 있는 셀이 3개 이상인 경우를 실제 헤더 행으로 판단 (제목 행 건너뛰기 위함)
            if sum(1 for cell in row if cell is not None) >= 3:
                header_row_index = i
                break
        
        print(f"유효 헤더 행 감지: {header_row_index}행")
        print("컬럼 헤더 확인:")
        header = [cell.value for cell in sheet[header_row_index]]
        for i, col_name in enumerate(header):
             if i < 28: # Print a reasonable number of headers
                print(f"\t{i}열: '{col_name}'")


        # Read data rows
        start_data_row = header_row_index + 1
        for row_index, row_cells in enumerate(sheet.iter_rows(min_row=start_data_row)):
            # Convert row_cells to a list of values
            row = [cell.value for cell in row_cells]

            # B열(1)이 비어있으면 건너뜀 (메시지의 시작점으로 간주)
            if len(row) <= 1 or not row[1]:
                continue

            current_row_num = start_data_row + row_index

            try:
                msg_id_hex = row[6]
                if msg_id_hex is None: continue
                
                # Validate that critical columns are numeric before processing the row
                msg_id = hex_to_dec(msg_id_hex)
                
                # Check if DLC and Cycle Time are numeric (suppress warnings for header rows)
                if not (str(row[8]).isdigit() or isinstance(row[8], int)) or \
                   not (str(row[9]).isdigit() or isinstance(row[9], int)):
                    continue

                # Detect CAN FD mode from Frame Type (H column, index 7)
                frame_type = row[7]
                if frame_type and "CAN FD" in str(frame_type):
                    self.use_canfd = True

                # New message found
                if msg_id not in self.messages:
                    sender_ecu = str(row[1]).strip()
                    self.validate_identifier(sender_ecu, "Send 제어기 (B열)", current_row_num)

                    source_addr = str(row[2]).strip()
                    
                    self.ecus.add(sender_ecu)
                    if source_addr:
                        self.ecu_addresses[sender_ecu] = source_addr

                    msg_name = str(row[5]).strip()
                    self.validate_identifier(msg_name, "Message Name (F열)", current_row_num)

                    self.messages[msg_id] = {
                        "name": msg_name,
                        "sender": sender_ecu,
                        "node_id": source_addr,
                        "dlc": row[8],
                        "cycle_time": row[9],
                        "cycle_time_fast": row[10],
                        "start_delay_time": row[12],
                        "tx_method": row[25],
                        "frame_type": row[7],
                        "nr_of_repetition": row[73],
                        "il_support": row[74],
                        "nm_message": row[75],
                        "diag_request": row[64],
                        "diag_response": row[65],
                        "tp_j1939_var_dlc_raw": row[76] if len(row) > 76 else None, # BY column
                        "signals": []
                    }
                
                # Add signal to the current message if signal name (N열, 13) exists
                if len(row) > 13 and row[13]:
                    sig_name = str(row[13]).strip()
                    self.validate_identifier(sig_name, "Signal Name (N열)", current_row_num)

                    if len(sig_name) > 32:
                        print(f"경고: 신호 이름 '{sig_name}'의 길이가 32자를 초과했습니다 ({len(sig_name)}자). (행 {current_row_num})", file=sys.stderr)

                    # Add receiver ECUs to the global ECU set
                    receivers_raw = row[16]
                    if receivers_raw:
                        receivers = [r.strip() for r in str(receivers_raw).split(',')]
                        for r in receivers:
                            self.validate_identifier(r, "Receiver ECU (Q열)", current_row_num)
                        self.ecus.update(receivers)
                    self.messages[msg_id]["signals"].append(row)

            except (IndexError, ValueError, TypeError) as e:
                print(f"경고: 행 {current_row_num} 처리 중 오류 발생, 건너뜁니다. 오류: {e}", file=sys.stderr)
                continue
        
        # 모드 출력 로직 개선
        mode_str = "Standard CAN"
        if self.use_canfd:
            mode_str = "CAN FD"
        elif any("J1939" in str(m['frame_type']) for m in self.messages.values()):
            mode_str = "J1939"
            
        print(f"감지된 모드: {mode_str}")
        print("\n메시지:")
        for msg_id, msg_data in self.messages.items():
             print(f"{msg_data['name']} (ID:{hex(msg_id)}), Sender={msg_data['sender']}, Node={msg_data['node_id']}, DLC={msg_data['dlc']}")

        print(f"\n{sum(len(m['signals']) for m in self.messages.values())}개 신호, {len(self.messages)}개 메시지")
        print("ECU Address 설정:")
        for ecu, addr in self.ecu_addresses.items():
            print(f"\t{ecu}: {addr}")


    def generate_dbc(self, output_path):
        """Generates and writes the DBC file content to the output path."""
        if not self.messages:
            print("오류: 생성할 메시지가 없습니다.", file=sys.stderr)
            return

        with open(output_path, 'w', encoding='utf-8') as f:
            # 1. Header
            f.write('VERSION ""\n\n')
            f.write('NS_ :\n\tNS_DESC_\n\tCM_\n\tBA_DEF_\n\tBA_\n\tVAL_\n\tCAT_DEF_\n\tCAT_\n\tFILTER\n\tBA_DEF_DEF_\n\tEV_DATA_\n\tENVVAR_DATA_\n\tSGTYPE_\n\tSGTYPE_VAL_\n\tBA_DEF_SGTYPE_\n\tBA_SGTYPE_\n\tSIG_TYPE_REF_\n\tVAL_TABLE_\n\tSIG_GROUP_\n\tSIG_VALTYPE_\n\tSIGTYPE_VALTYPE_\n\n')
            f.write('BS_:\n\n')

            # 2. ECUs (BU_)
            f.write(f'BU_: {" ".join(sorted(list(self.ecus)))}\n\n')

            # 3. Messages (BO_) and Signals (SG_)
            for msg_id, msg in sorted(self.messages.items()):
                # I열(DLC) 값을 엑셀 그대로 반영 (8로 제한하는 로직 제거)
                dlc = int(msg['dlc'])
                
                # Extended CAN ID (29bit) 처리: MSB(Bit 31)를 1로 설정 (DBC 표준)
                dbc_id = msg_id | 0x80000000 if msg_id > 0x7FF else msg_id
                f.write(f'BO_ {dbc_id} {msg["name"]}: {dlc} {msg["sender"]}\n')
                
                for sig_row in msg["signals"]:
                    try:
                        name = sig_row[13]
                        start = int(sig_row[20])
                        length = int(sig_row[18])
                        byte_order = '0' if 'Motorola' in str(sig_row[21]) else '1' # 0=big, 1=little
                        sign = '+' if 'Unsigned' in str(sig_row[40]) else '-'
                        scale = sig_row[44] if sig_row[44] is not None else 1.0
                        offset = sig_row[45] if sig_row[45] is not None else 0.0
                        min_val = sig_row[46] if sig_row[46] is not None else 0
                        max_val = sig_row[47] if sig_row[47] is not None else (2**length) * scale - (1 if sign == '+' else 0)
                        unit = sig_row[49] or ""
                        receivers_raw = sig_row[16]
                        receivers = [r.strip() for r in str(receivers_raw).split(',')] if receivers_raw else ['Vector__XXX']
                        
                        f.write(f' SG_ {name} : {start}|{length}@{byte_order}{sign} ({scale},{offset}) [{min_val}|{max_val}] "{unit}" {",".join(receivers)}\n')
                    except (IndexError, TypeError, ValueError) as e:
                        print(f"경고: 신호 처리 중 오류 (메시지: {msg['name']}), 건너뜁니다. 오류: {e}", file=sys.stderr)
                        continue
                f.write('\n')

            # 4. Comments (CM_)
            for msg_id, msg in self.messages.items():
                dbc_id = msg_id | 0x80000000 if msg_id > 0x7FF else msg_id
                for sig_row in msg["signals"]:
                    desc = sig_row[15]
                    if desc:
                        f.write(f'CM_ SG_ {dbc_id} {sig_row[13]} "{desc}";\n')
            f.write('\n')

            # 5. Attribute Definitions (BA_DEF_)
            f.write('BA_DEF_ "BusType" STRING;\n')
            f.write('BA_DEF_ BU_ "NmStationAddress" HEX 0 255;\n')
            f.write('BA_DEF_ BO_ "VFrameFormat" ENUM "StandardCAN","ExtendedCAN","reserved","J1939PG","reserved","reserved","reserved","reserved","reserved","reserved","reserved","reserved","reserved","reserved","StandardCAN_FD","ExtendedCAN_FD";\n')
            f.write('BA_DEF_ BO_ "GenMsgCycleTime" INT 0 200000;\n')
            f.write('BA_DEF_ BO_ "GenMsgCycleTimeFast" INT 0 200000;\n')
            f.write('BA_DEF_ BO_ "GenMsgStartDelayTime" INT 0 200000;\n')
            f.write('BA_DEF_ BO_ "GenMsgSendType" ENUM "Cyclic","NoMsgSendType";\n')
            f.write('BA_DEF_ BO_ "GenMsgNrOfRepetition" INT 0 100;\n')
            f.write('BA_DEF_ BO_ "TpJ1939VarDlc" ENUM "No","Yes";\n')
            f.write('BA_DEF_ BO_ "GenMsgDelayTime" INT 0 1000;\n')
            f.write('BA_DEF_ BO_ "GenMsgRequestable" INT 0 1;\n')
            f.write('BA_DEF_ BO_ "GenMsgILSupport" ENUM "No","Yes";\n')
            f.write('BA_DEF_ BO_ "NmMessage" ENUM "No","Yes";\n')
            f.write('BA_DEF_ BO_ "DiagRequest" ENUM "No","Yes";\n')
            f.write('BA_DEF_ BO_ "DiagResponse" ENUM "No","Yes";\n')
            f.write('BA_DEF_ SG_ "GenSigSendType" ENUM "NotUsed","OnWrite","OnWriteWithRepetition","OnChange","OnChangeWithRepetition","IfActive","IfActiveWithRepetition","NoSigSendType";\n')
            f.write('BA_DEF_ SG_ "GenSigStartValue" INT 0 100000;\n\n')

            # 6. Default Attribute Values (BA_DEF_DEF_)
            bus_type_def = "CAN FD" if self.use_canfd else "CAN"
            f.write(f'BA_DEF_DEF_ "BusType" "{bus_type_def}";\n')
            f.write('BA_DEF_DEF_ "NmStationAddress" 0;\n')
            default_frame_format = "StandardCAN_FD" if self.use_canfd else "StandardCAN"
            f.write(f'BA_DEF_DEF_ "VFrameFormat" "{default_frame_format}";\n')
            f.write('BA_DEF_DEF_ "GenMsgCycleTime" 0;\n')
            f.write('BA_DEF_DEF_ "GenMsgCycleTimeFast" 0;\n')
            f.write('BA_DEF_DEF_ "GenMsgStartDelayTime" 0;\n')
            f.write('BA_DEF_DEF_ "GenMsgSendType" "NoMsgSendType";\n')
            f.write('BA_DEF_DEF_ "GenMsgNrOfRepetition" 0;\n')
            f.write('BA_DEF_DEF_ "TpJ1939VarDlc" "No";\n')
            f.write('BA_DEF_DEF_ "GenMsgDelayTime" 0;\n')
            f.write('BA_DEF_DEF_ "GenMsgRequestable" 0;\n')
            f.write('BA_DEF_DEF_ "GenMsgILSupport" "Yes";\n')
            f.write('BA_DEF_DEF_ "NmMessage" "No";\n')
            f.write('BA_DEF_DEF_ "DiagRequest" "No";\n')
            f.write('BA_DEF_DEF_ "DiagResponse" "No";\n')
            f.write('BA_DEF_DEF_ "GenSigSendType" "NoSigSendType";\n')
            f.write('BA_DEF_DEF_ "GenSigStartValue" 0;\n\n')

            # 7. Attribute Values (BA_)
            f.write(f'BA_ "BusType" "{bus_type_def}";\n')

            # ECU Attributes
            for ecu, addr in self.ecu_addresses.items():
                f.write(f'BA_ "NmStationAddress" BU_ {ecu} {hex_to_dec(addr)};\n')
            
            # Message and Signal Attributes
            print("\n메시지 속성 설정:")
            for msg_id, msg in self.messages.items():
                dbc_id = msg_id | 0x80000000 if msg_id > 0x7FF else msg_id

                # VFrameFormat
                frame_type = str(msg['frame_type']) if msg['frame_type'] else ""
                is_extended = msg_id > 0x7FF or 'Extended' in frame_type
                
                if "J1939" in frame_type:
                    val = 3
                elif self.use_canfd or "CAN FD" in frame_type:
                    val = 15 if is_extended else 14
                else: # Fallback for classic CAN
                    val = 1 if is_extended else 0

                f.write(f'BA_ "VFrameFormat" BO_ {dbc_id} {val};\n')

                # Timing
                if msg['cycle_time'] is not None:
                    f.write(f'BA_ "GenMsgCycleTime" BO_ {dbc_id} {int(msg["cycle_time"])};\n')
                if msg['cycle_time_fast'] is not None:
                    f.write(f'BA_ "GenMsgCycleTimeFast" BO_ {dbc_id} {int(msg["cycle_time_fast"])};\n')
                if msg['start_delay_time'] is not None:
                     f.write(f'BA_ "GenMsgStartDelayTime" BO_ {dbc_id} {int(msg["start_delay_time"])};\n')
                
                # Send Type
                send_type_map = {"Cyclic": 0, "NoMsgSendType": 1}
                send_type_val = send_type_map.get(msg['tx_method'])
                if send_type_val is not None:
                    f.write(f'BA_ "GenMsgSendType" BO_ {dbc_id} {send_type_val};\n')

                # Repetition
                if msg['nr_of_repetition'] is not None and int(msg['nr_of_repetition']) != 0:
                    f.write(f'BA_ "GenMsgNrOfRepetition" BO_ {dbc_id} {int(msg["nr_of_repetition"])};\n')
                
                # TpJ1939VarDlc
                tp_val = None
                if 'DM1_' in msg['name']:
                    tp_val = 1 # Yes
                elif msg.get('tp_j1939_var_dlc_raw'):
                    raw_val = str(msg['tp_j1939_var_dlc_raw']).lower()
                    if raw_val == 'yes':
                        tp_val = 1 # Yes
                    elif raw_val == 'no':
                        tp_val = 0 # No
                if tp_val is not None:
                    f.write(f'BA_ "TpJ1939VarDlc" BO_ {dbc_id} {tp_val};\n')
                
                # Fixed values
                f.write(f'BA_ "GenMsgDelayTime" BO_ {dbc_id} 0;\n')
                f.write(f'BA_ "GenMsgRequestable" BO_ {dbc_id} 1;\n')

                # Values only if not default
                if str(msg.get('il_support')).lower() == 'no':
                     f.write(f'BA_ "GenMsgILSupport" BO_ {dbc_id} 0;\n') # 0=No
                if str(msg.get('nm_message')).lower() == 'yes':
                     f.write(f'BA_ "NmMessage" BO_ {dbc_id} 1;\n') # 1=Yes
                if str(msg.get('diag_request')).lower() == 'yes':
                     f.write(f'BA_ "DiagRequest" BO_ {dbc_id} 1;\n') # 1=Yes
                if str(msg.get('diag_response')).lower() == 'yes':
                     f.write(f'BA_ "DiagResponse" BO_ {dbc_id} 1;\n') # 1=Yes

                # Signal attributes
                for sig_row in msg["signals"]:
                    sig_name = sig_row[13]
                    # GenSigSendType
                    send_type_str = sig_row[27]
                    sig_send_type_map = {"NotUsed":0, "OnWrite":1, "OnWriteWithRepetition":2, "OnChange":3, "OnChangeWithRepetition":4, "IfActive":5, "IfActiveWithRepetition":6, "NoSigSendType":7}
                    sig_send_val = sig_send_type_map.get(send_type_str)
                    if sig_send_val is not None and sig_send_val != 7: # Don't write default
                        f.write(f'BA_ "GenSigSendType" SG_ {dbc_id} {sig_name} {sig_send_val};\n')

                    # GenSigStartValue
                    init_val_raw = sig_row[41]
                    if init_val_raw is not None:
                        init_val = hex_to_dec(init_val_raw)
                        f.write(f'BA_ "GenSigStartValue" SG_ {dbc_id} {sig_name} {init_val};\n')

        print(f"\nDBC 생성 완료: {output_path}")


if __name__ == "__main__":
    class TextRedirector:
        def __init__(self, widget, tag="stdout"):
            self.widget = widget
            self.tag = tag

        def write(self, str):
            self.widget.configure(state="normal")
            self.widget.insert("end", str, (self.tag,))
            self.widget.see("end")
            self.widget.configure(state="disabled")
            self.widget.update()

        def flush(self):
            pass

    class DBCConverterApp:
        def __init__(self, root):
            self.root = root
            self.root.title("Excel to CAN FD DBC Converter")
            self.root.geometry("600x600")

            self.excel_path = None
            self.sheet_vars = {}

            # UI Elements
            self.lbl_title = tk.Label(root, text="Excel to CAN FD DBC Converter", font=("Arial", 16, "bold"), bg="dark green", fg="white")
            self.lbl_title.pack(fill='x', pady=(0, 5))

            frame_top = tk.Frame(root)
            frame_top.pack(pady=10, fill='x', padx=10)

            self.btn_open = tk.Button(frame_top, text="Excel 파일 열기", command=self.open_file)
            self.btn_open.pack(side='left')

            self.lbl_file = tk.Label(frame_top, text="파일이 선택되지 않았습니다.", fg="gray")
            self.lbl_file.pack(side='left', padx=10)

            frame_opts = tk.Frame(root)
            frame_opts.pack(pady=5, fill='x', padx=10)

            tk.Label(frame_opts, text="System명:").pack(side='left')
            self.entry_sys = tk.Entry(frame_opts, width=12)
            self.entry_sys.insert(0, "MySys")
            self.entry_sys.pack(side='left', padx=5)

            self.var_date = tk.BooleanVar(value=True)
            tk.Checkbutton(frame_opts, text="날짜", variable=self.var_date).pack(side='left', padx=5)

            tk.Label(frame_opts, text="Ver:").pack(side='left')
            self.entry_ver = tk.Entry(frame_opts, width=8)
            self.entry_ver.insert(0, "v00")
            self.entry_ver.pack(side='left', padx=5)

            tk.Label(root, text="변환할 시트 선택:").pack(anchor='w', padx=10, pady=(10, 0))

            self.frame_sheets = tk.Frame(root)
            self.frame_sheets.pack(pady=5, fill='both', expand=True, padx=10)

            self.btn_convert = tk.Button(root, text="DBC 변환 시작", command=self.convert, state=tk.DISABLED, height=2, font=("Arial", 12, "bold"), bg="blue", fg="white")
            self.btn_convert.pack(pady=20, fill='x', padx=20)

            # Log Window
            tk.Label(root, text="로그:").pack(anchor='w', padx=10)
            self.log_text = scrolledtext.ScrolledText(root, state='disabled', height=15)
            self.log_text.pack(fill='both', expand=True, padx=10, pady=(0, 10))
            self.log_text.tag_config("stdout", foreground="black")
            self.log_text.tag_config("stderr", foreground="red")

            sys.stdout = TextRedirector(self.log_text, "stdout")
            sys.stderr = TextRedirector(self.log_text, "stderr")

        def open_file(self):
            path = filedialog.askopenfilename(
                title="Excel 파일 선택",
                initialdir=r"C:\Python\Excel2CANFD",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
            )
            if path:
                self.excel_path = path
                self.lbl_file.config(text=Path(path).name, fg="black")
                self.load_sheets()

        def load_sheets(self):
            for widget in self.frame_sheets.winfo_children():
                widget.destroy()
            self.sheet_vars.clear()

            try:
                wb = openpyxl.load_workbook(self.excel_path, read_only=True)
                for sheet in wb.sheetnames:
                    var = tk.BooleanVar()
                    chk = tk.Checkbutton(self.frame_sheets, text=sheet, variable=var)
                    chk.pack(anchor='w')
                    self.sheet_vars[sheet] = var
                wb.close()
                self.btn_convert.config(state=tk.NORMAL)
            except Exception as e:
                messagebox.showerror("오류", f"파일을 읽을 수 없습니다: {e}")

        def convert(self):
            selected = [name for name, var in self.sheet_vars.items() if var.get()]
            if not selected:
                messagebox.showwarning("경고", "변환할 시트를 하나 이상 선택하세요.")
                return

            sys_name = self.entry_sys.get().strip()
            ver_str = self.entry_ver.get().strip()
            date_str = datetime.datetime.now().strftime("%Y%m%d") if self.var_date.get() else ""

            self.log_text.configure(state='normal')
            self.log_text.delete(1.0, tk.END)
            self.log_text.configure(state='disabled')
            print("변환 작업을 시작합니다...")

            count = 0
            for sheet_name in selected:
                try:
                    # Filename: System_Sheet_Date_Ver.dbc
                    parts = [p for p in [sys_name, sheet_name, date_str, ver_str] if p]
                    output_filename = "_".join(parts) + ".dbc"
                    output_path = Path(self.excel_path).parent / output_filename
                    gen = CANDBCGenerator(self.excel_path, sheet_name)
                    gen.parse_signals()
                    gen.generate_dbc(output_path)
                    count += 1
                    # print(f"변환 완료: {output_path}")
                except Exception as e:
                    print(f"\n[오류] '{sheet_name}' 변환 중 문제 발생:\n{e}", file=sys.stderr)
            
            if count > 0:
                print(f"\n[완료] 총 {count}개의 DBC 파일이 생성되었습니다.")
            else:
                print("\n[알림] 생성된 파일이 없습니다.")

    root = tk.Tk()
    app = DBCConverterApp(root)
    root.mainloop()
